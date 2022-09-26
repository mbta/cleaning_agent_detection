import dataclasses
import re
import os
import datetime
import dataclasses
from dataclasses import dataclass
from dataclasses import field
from typing import List
from typing import Dict
from typing import Any
from typing import Tuple
from typing import Union
from typing import Iterator
import pathlib
import logging

import csv
from openpyxl import load_workbook
from openpyxl import Workbook


def check_file_path(filename: str) -> Tuple[bool, pathlib.Path]:
    """
    check that file exists and extension is readable by script
    only .csv or .xlsx files accepted

    return: 
        valid_file: bool indicating file validity
        input_file: pathlib.Path for file
    """
    input_file = pathlib.Path(filename)
    valid_file = True
    if not input_file.is_file():
        logging.error(f"{input_file} does not exist.")
        valid_file = False
    elif input_file.suffix.lower() not in ('.xlsx', '.csv',):
        logging.error(f"Please input path to '.csv' or '.xlsx' files.")
        valid_file = False
    return (valid_file, input_file)


@dataclass
class CleanRecord:
    """
    holder dataclass for cleaning / maintenance records
    """
    id: str
    location: str
    zone: str
    title: str
    dt: datetime.datetime 
    elevator: str = None
    has_alert: bool = False

    def __post_init__(self):
        self.str_to_datetime()
        self.clean_zone()

    def str_to_datetime(self):
        """
        convert str to datetime, if needed
        """
        strptime_format = "%m/%d/%Y %H:%M"
        if isinstance(self.dt, datetime.datetime):
            return
        self.dt = datetime.datetime.strptime(self.dt, strptime_format)
    
    def clean_zone(self):
        """
        clean 'zone' field, strip numbers, spaces special characters and make
        all lower case
        """
        self.zone =  re.sub(r"[\W\d]","",self.zone).lower()


def make_clean_record(record: List[str], column_lookups: Dict[str,int]):
    """
    create CleanRecord object from spreadsheet row
    """
    return CleanRecord(
        id=str(record[column_lookups["#"]]),
        location=str(record[column_lookups["Address"]]),
        zone=str(record[column_lookups["Zone"]]),
        title=str(record[column_lookups["Title"]]),
        dt=record[column_lookups["Created"]],
    )


@dataclass
class AlertRecord:
    """
    holder dataclass for sensor alert records
    """
    dt: datetime.datetime 
    location: str
    id: str
    status: str
    zone: str = None
    elevator: str = None

    def __post_init__(self):
        self.str_to_datetime()
        self.get_elevator_from_location()
        self.get_zone_from_location()

    def str_to_datetime(self):
        """
        convert str to datetime, if needed
        """
        strptime_format = "%Y-%m-%d %H:%M:%S"
        if isinstance(self.dt, datetime.datetime):
            return
        self.dt = datetime.datetime.strptime(self.dt, strptime_format)

    def get_elevator_from_location(self):
        """
        pull elevator number from 'location' 
        """
        match_obj = re.search(r"\d{3,}", self.location)
        if match_obj is None:
            raise TypeError(f"No elevator number found in location: {self.location}")
        self.elevator = match_obj.group()

    def get_zone_from_location(self):
        """
        clean 'zone' field, strip numbers, spaces special characters and make
        all lower case
        """
        self.zone = re.sub(r"[\W\d]","",self.location).lower()
        if self.zone == "downtowncrossing":
            self.zone = "dtx"


@dataclass
class Metric:
    """
    holder dataclass for metrics information
    """
    true_positive: List[Tuple[AlertRecord, CleanRecord]] = field(default_factory=list)
    false_positive: List[AlertRecord] = field(default_factory=list)
    false_negative: List[CleanRecord] = field(default_factory=list)

    def __str__(self):
        label_len = 20
        num_len = 10
        row_len = label_len + num_len + 5
        return (
            f"{'*'*row_len}\n"
            f"*{str.center('Results',row_len-2)}*\n"
            f"{'*'*row_len}\n"
            f"* {str.ljust('True Positive',label_len)}* {str.ljust(str(len(self.true_positive)),num_len)}*\n"
            f"* {str.ljust('False Positive',label_len)}* {str.ljust(str(len(self.false_positive)),num_len)}*\n"
            f"* {str.ljust('False Negative',label_len)}* {str.ljust(str(len(self.false_negative)),num_len)}*\n"
            f"{'*'*row_len}\n"
        )

    def get_table(self):
        return [
            ("Metric", "Count"),
            ("True Positive", len(self.true_positive)),
            ("False Positive", len(self.false_positive)),
            ("False Negative", len(self.false_negative)),
        ]

    def true_positive_table(self):
        return_table = [("Alert_Time","Cleaning_Time","Alert_Location","Cleaning_Location","Alert_ID","Cleaning_ID","Cleaning_Title","Zone","Elevator"),]
        for record in self.true_positive:
            alert = record[0]
            clean = record[1]
            return_table.append((
                alert.dt,
                clean.dt,
                alert.location,
                clean.location,
                alert.id,
                clean.id,
                clean.title,
                alert.zone,
                alert.elevator,
            ))
        return return_table

    def false_positive_table(self):
        return_table = [("Time","Location","ID","Zone","Elevator"),]
        for record in self.false_positive:
            return_table.append((
                record.dt,
                record.location,
                record.id,
                record.zone,
                record.elevator,
            ))
        return return_table

    def false_negative_table(self):
        return_table = [("Time","ID","Location","Title","Zone","Elevator"),]
        for record in self.false_negative:
            return_table.append((
                record.dt,
                record.id,
                record.location,
                record.title,
                record.zone,
                record.elevator,
            ))
        return return_table




def make_alert_record(record: List[str], column_lookups: Dict[str,int]):
    """
    create AlertRecord object from spreadsheet row
    """
    return AlertRecord(
        dt=record[column_lookups["Date & Time Stamp"]],
        location=str(record[column_lookups["Location Elevator #"]]),
        id=str(record[column_lookups["Alert ID"]]),
        status=str(record[column_lookups["Status"]]),
    )


def csv_as_list(filename: str) -> List[List[str]]:
    """
    read csv file and return data as list of lists
    """
    with open(filename, newline="") as read_file:
        csv_reader = csv.reader(read_file)
        return [row for row in csv_reader]


def xlsx_as_dict_of_lists(filename: str) -> Dict[str, List[List[Any]]]:
    """
    read Excel file and return all sheets as dictionary object
    dictionary keys are sheet names
    dictionary values are list of list with sheet data
    """
    return_dict = {}
    wb = load_workbook(filename=filename, read_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        return_dict[sheet] = [[cell.value for cell in row] for row in ws.rows]

    return return_dict


def get_header_index(headers_to_find: List[str], header_values: List[str]) -> Dict[str,int]:
    """
    for list of header names, and list of possible header values
    return dictionary of header names
    dictionary keys are header names
    dictionary values are index of header columns in list
    """
    header_diff = set(headers_to_find) - set(header_values)
    if len(header_diff) > 0:
        raise KeyError(f"Required headers missing from file: {header_diff}")

    return_dict = {}
    for required_header in headers_to_find:
        for column_number, header in enumerate(header_values):
            if header == required_header:
                return_dict[required_header] = column_number
                break

    return return_dict


def get_elevator_records(record: CleanRecord) -> Iterator[CleanRecord]:
    """
    explode cleaning record into invidual records if multiple elevators are 
    identified in location field
    """
    # if "elev" in record.location.lower():
    elevator_numbers = re.findall(r"\d{3}", record.location)
    # if no elevators found, empty list returned no records added to output
    for elevator_number in elevator_numbers:
        yield dataclasses.replace(record, elevator=elevator_number)


def check_data_headers(data: Union[list,dict], required_cols: List[str]):
    """
    data provided as either list of lists (csv file) 
    OR dict containing list of lists (xlsx file)

    this will scan the first row of each file for required header columns and
    return a mapping dictionary between required columns and column numbers
    """
    headers_found = False
    columns_lookup = None
    # scan csv file data
    if isinstance(data, list):
        try:
            columns_lookup = get_header_index(required_cols, data[0])
            headers_found = True
        except Exception as e:
            logging.error(e)
    # scan xlsx file data
    else:
        for sheet_data in data.values():
            try:
                columns_lookup = get_header_index(required_cols, sheet_data[0])
            except Exception as e:
                logging.error(e)
                continue
            else:
                data = sheet_data
                headers_found = True
                break
    if headers_found:
        logging.info("All required headers found in file!")
    return (headers_found, columns_lookup, data)


def get_file_path(file_type: str, required_cols: List[str], filename = None):
    """
    input file report path and check for valid data
    """
    valid_file = False
    while valid_file is False:
        print(f"Enter path for {file_type} file:")
        if filename is None:
            input_file = input()
        else:
            input_file = filename
        (valid_file, input_file) = check_file_path(input_file)
        if valid_file is False:
            continue

        logging.info(f"Loading data from {input_file}...")
        if input_file.suffix.lower() == ".csv":
            data = csv_as_list(input_file)
        elif input_file.suffix.lower() == ".xlsx":
            data = xlsx_as_dict_of_lists(input_file)
        else:
            raise TypeError(f"file type not expected")
        logging.info(f"Data loaded from {input_file}")
        (valid_file, columns_lookup, data) = check_data_headers(data, required_cols)


    return (data, columns_lookup)

def list_to_dict(list_data: List[Any], dt_min: datetime.datetime, dt_max: datetime.datetime) -> Dict[str,Dict[str,Any]]:
    """
    take spreadsheet data, list of records, and transform to dict of dicts of list of records
    records will be limited to datetimes between dt_min and dt_max

    dictionary structure:
        Zone: Dict[Elevator: List[records]]

    nested list of records will be sorted by datetime field
    """
    return_dict = {}
    for record in list_data:
        if record.dt < dt_min or record.dt > dt_max:
            continue
        if record.zone not in return_dict:
            return_dict[record.zone] = {}
        if record.elevator not in return_dict[record.zone]:
            return_dict[record.zone][record.elevator] = []
        return_dict[record.zone][record.elevator].append(record)

    for zone, zone_data in return_dict.items():
        for elevator, elevator_data in zone_data.items():
            return_dict[zone][elevator] = sorted(elevator_data, key=lambda x: x.dt)
    return return_dict


def pull_sense_file() -> Tuple[List[AlertRecord], datetime.datetime, datetime.datetime]:
    """
    ask for sensor data file path 

    process extract and process sensor data file, limiting to "cleaning" records

    return list of records and min and max datetime of records
    """
    required_sense_cols = [
        "Date & Time Stamp",
        "Location Elevator #",
        "Alert ID",
        "Status",
    ]
    (data, columns_lookup) = get_file_path("Sensitics report", required_sense_cols)
    data = [make_alert_record(record, columns_lookup) for record in data[1:]]
    
    # limit alert records to "cleaning" records
    data = [record for record in data if "clean" in record.status.lower()]
    min_dt = data[0].dt
    max_dt = data[0].dt
    # remove triplicate alerts
    new_data = [data[0],]
    for record in data[1:]:
        time_diff = (record.dt - new_data[-1].dt).total_seconds()
        # do not include alert if pervious alert has same id
        # and occured within the last 6 minutes
        if record.id == new_data[-1].id and time_diff <= 360 and time_diff > 0:
            continue

        if record.dt < min_dt:
            min_dt = record.dt
        if record.dt > max_dt:
            max_dt = record.dt
        new_data.append(record)

    if len(new_data) > 0:
        logging.info(f"{len(new_data)} cleaning sensor records found from {min_dt.strftime('%b %d, %Y')} to {max_dt.strftime('%b %d, %Y')}")
    
    return (new_data, min_dt, max_dt)

def pull_clean_report_file(sensor_locations) -> Tuple[List[CleanRecord], datetime.datetime, datetime.datetime]:
    """
    ask for maintenance log data file path 

    process extract and cleaning records limiting records to zone and elevator numbers
    found in sensor alert data

    return list of records and min and max datetime of records
    """
    required_clean_columns = [
        "#",
        "Title",
        "Address",
        "Created",
        "Zone",
        ]
    (data, columns_lookup) = get_file_path("maintenance request", required_clean_columns)
    filtered_data = []
    for record in data[1:]:
        # skip records with no valid data
        if len(record) == record.count(None):
            continue
        
        record = make_clean_record(record, columns_lookup)
        for new_record in get_elevator_records(record):
            # limit cleaning records to zones and elevators numbers found in
            # alerts data
            if (new_record.zone, new_record.elevator) in sensor_locations:
                # get elevator cleaning records and explode records containing reference
                # to more than one elevator
                filtered_data.append(new_record)

    # get min and max cleaning data datetimes
    min_dt = datetime.datetime(year=3000,month=1, day=1)
    max_dt = datetime.datetime(year=1,month=1, day=1)
    for record in filtered_data:
            if record.dt < min_dt:
                min_dt = record.dt
            if record.dt > max_dt:
                max_dt = record.dt
    if len(filtered_data) > 0:
        logging.info(f"{len(filtered_data)} maintenance cleaning records found from {min_dt.strftime('%b %d, %Y')} to {max_dt.strftime('%b %d, %Y')}")

    return (filtered_data, min_dt, max_dt)


def export_metrics(metrics: Metric) -> None:
    """
    if metrics counts exist, export metrics as csv file to home directory
    """
    if len(metrics.true_positive) + len(metrics.false_negative) + len(metrics.false_positive) == 0:
        logging.info(f"No results to export.")
        return

    filename = f"detect_cleaning_agent_metrics_{datetime.datetime.now().isoformat()}.xlsx"
    file_path = os.path.join(pathlib.Path.home(), filename)
    wb = Workbook()
    ws = wb.create_sheet(title="Metrics")
    for row in metrics.get_table():
        ws.append(row)

    ws = wb.create_sheet(title="True_Positives")
    for row in metrics.true_positive_table():
        ws.append(row)

    ws = wb.create_sheet(title="False_Positives")
    for row in metrics.false_positive_table():
        ws.append(row)

    ws = wb.create_sheet(title="False_Negatives")
    for row in metrics.false_negative_table():
        ws.append(row)

    wb.remove(wb['Sheet'])
    wb.save(filename=file_path)
    logging.info(f"Wrote results file to: {file_path}")
